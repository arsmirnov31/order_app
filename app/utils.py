import openpyxl
import re
from decimal import Decimal
import decimal
from .db import get_db_connection
import psycopg2.extras
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Border, Side
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


def generate_excel(preview_data, points, report_date):

    thin = Side(style="thin")
    medium = Side(style="medium")

    thin_border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    medium_border = Border(
        left=medium,
        right=medium,
        top=medium,
        bottom=medium
    )

    wb = Workbook()
    wb.remove(wb.active)

    # стили
    header_font = Font(bold=True)
    align_left_wrap = Alignment(horizontal="left", wrap_text=True)
    align_center_wrap = Alignment(horizontal="center", wrap_text=True)
    align_right = Alignment(horizontal="right")

    for category in preview_data:

        # Excel ограничение 31 символ
        sheet_name = (category["category_name"] or "Без категории")[:31]

        ws = wb.create_sheet(title=sheet_name)

        # -------------------------
        # настройки страницы
        # -------------------------
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.fitToPage = True

        ws.page_margins = PageMargins(
            left=0.3,
            right=0.3,
            top=0.5,
            bottom=0.5
        )

        # -------------------------
        # дата отчета
        # -------------------------
        ws.cell(row=1, column=1, value=f"Дата: {report_date}").font = Font(bold=True)

        # -------------------------
        # заголовок таблицы
        # -------------------------
        header_row = 3

        cell = ws.cell(row=header_row, column=1, value="Ассортимент")
        cell.font = header_font
        cell.alignment = align_center_wrap
        cell.border = medium_border

        for col_idx, point in enumerate(points, start=2):
            cell = ws.cell(row=header_row, column=col_idx, value=point["name"])
            cell.font = header_font
            cell.alignment = align_center_wrap
            cell.border = thin_border

        last_col = len(points) + 2

        cell = ws.cell(row=header_row, column=last_col, value="Итого")
        cell.font = header_font
        cell.alignment = align_center_wrap
        cell.border = thin_border

        # повтор заголовка при печати
        ws.print_title_rows = f"{header_row}:{header_row}"

        # -------------------------
        # данные
        # -------------------------
        row_idx = header_row + 1

        for row in category["rows"]:

            # название продукта
            cell = ws.cell(row=row_idx, column=1, value=row["product_name"])
            cell.alignment = align_left_wrap
            cell.border = medium_border
            # значения
            for col_idx, val in enumerate(row["point_values"], start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = align_right
                cell.border = thin_border

            # итог
            cell = ws.cell(row=row_idx, column=last_col, value=row["total_display"])
            cell.alignment = align_right
            cell.font = Font(bold=True)
            cell.border = thin_border

            row_idx += 1

        # -------------------------
        # ширина колонок
        # -------------------------
        ws.column_dimensions["A"].width = 35

        for i in range(2, last_col + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 12

    # -------------------------
    # сохраняем в память
    # -------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def format_date(d):
    if isinstance(d, str):
        return d
    return d.strftime("%d.%m.%Y")

def build_report_data(report_date, all_points, selected_point_ids, excel_data=None):

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # -------------------------
    # точки
    # -------------------------
    cur.execute("""
        select point_id, name
        from points
        where is_active = true
        order by sort_order
    """)
    points = cur.fetchall()

    # -------------------------
    # выбор точек
    # -------------------------
    if all_points:
        selected_point_ids = [p["point_id"] for p in points]

    no_points_selected = not all_points and not selected_point_ids

    rows = []

    if not no_points_selected:

        cur.execute("""
            select
                pc.name  as product_category_name,
                p.product_id,
                p.name   as product_name,
                u.name   as unit_name,
                pt.point_id,
                pt.name  as point_name,
                oi.quantity
            from orders o
            join order_items oi on oi.order_id = o.order_id
            join products p on p.product_id = oi.product_id
            left join product_categories pc on pc.product_category_id = p.product_category_id
            left join public.unit_of_measure u on u.measure_id = p.measure_id
            join points pt on pt.point_id = o.point_id
            where
                o.status_id = 4
                and o.order_date::date = %(report_date)s
                and (
                    %(all_points)s = true
                    or o.point_id = any(%(point_ids)s)
                )
            order by
                pt.sort_order nulls last,
                pc.sort_order,
                p.sort_order
        """, {
            "report_date": report_date,
            "all_points": all_points,
            "point_ids": selected_point_ids
        })

        rows = cur.fetchall()

    # -------------------------
    # сборка data
    # -------------------------
    data = {}

    for row in rows:

        category_name = row["product_category_name"] or "Без категории"
        product_id = row["product_id"]

        if category_name not in data:
            data[category_name] = {}

        if product_id not in data[category_name]:
            data[category_name][product_id] = {
                "product_name": row["product_name"],
                "unit_name": row["unit_name"],
                "normalized_name": normalize_product_name(row["product_name"]),
                "points": {}
            }

        data[category_name][product_id]["points"][row["point_id"]] = row["quantity"]

    # -------------------------
    # merge
    # -------------------------
    if excel_data:
        data = merge_with_excel(data, excel_data, points)

    # -------------------------
    # добиваем нулями
    # -------------------------
    point_order = [p["point_id"] for p in points]

    for category in data.values():
        for product in category.values():
            for point_id in point_order:
                if point_id not in product["points"]:
                    product["points"][point_id] = 0

    # -------------------------
    # форматирование
    # -------------------------
    from decimal import Decimal

    def format_quantity(value, unit_name):
        if value is None:
            value = Decimal(0)

        if unit_name == "кг":
            return f"{value:.3f}"
        return str(int(round(value)))

    # -------------------------
    # preview
    # -------------------------
    preview_data = []

    for category_name, products in data.items():

        category_block = {
            "category_name": category_name,
            "rows": []
        }

        for product in products.values():

            unit_name = product["unit_name"]

            row = {
                "product_name": f'{product["product_name"]} ({unit_name})',
                "point_values": [],
                "total_display": ""
            }

            total = 0

            for point_id in point_order:
                val = product["points"].get(point_id, 0)

                row["point_values"].append(
                    format_quantity(val, unit_name)
                )

                total += val

            row["total_display"] = format_quantity(total, unit_name)

            category_block["rows"].append(row)

        preview_data.append(category_block)

    cur.close()
    conn.close()

    return preview_data, points


def parse_excel(file):

    wb = openpyxl.load_workbook(file, data_only=True)  
    # data_only=True → берём значения, а не формулы

    sheets_data = {}

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        sheets_data[normalize_category(sheet_name)] = parse_sheet(ws)

    return sheets_data


def normalize_product_name(name: str) -> str:
    if not name:
        return ""

    name = name.lower().strip()

    # убираем всё в скобках
    name = re.sub(r"\(.*?\)", "", name)

    # убираем лишние пробелы
    name = re.sub(r"\s+", " ", name)

    return name.strip()


def parse_sheet(ws):

    HEADER_ROW = 3
    PRODUCT_COL = 2  # B

    data = {}

    # -------------------------
    # 1. читаем заголовки (точки)
    # -------------------------
    headers = []
    col_idx = PRODUCT_COL + 1  # начинаем с C

    while True:
        value = ws.cell(row=HEADER_ROW, column=col_idx).value

        # стоп, если дошли до пустой или "Итого"
        if value is None or str(value).strip().lower() == "итого":
            break

        headers.append(str(value).strip())
        col_idx += 1

    # -------------------------
    # 2. читаем строки
    # -------------------------
    row_idx = HEADER_ROW + 1

    while True:

        product_name = ws.cell(row=row_idx, column=PRODUCT_COL).value

        # стоп, если пустая строка
        if product_name is None:
            break

        product_name = str(product_name).strip()

        if not product_name:
            row_idx += 1
            continue

        normalized_name = normalize_product_name(product_name)

        values = {}

        for i, point_name in enumerate(headers):
            col_idx = PRODUCT_COL + 1 + i

            val = ws.cell(row=row_idx, column=col_idx).value
            val = safe_decimal(val)

            values[point_name] = val

        data[normalized_name] = values

        row_idx += 1

    return data

def normalize_category(name):
    return str(name).strip().lower()

def merge_with_excel(db_data, excel_data, points):
    point_name_to_id = {p["name"]: p["point_id"] for p in points}

    print(f"Пытаемся найти точку {point_name_to_id}")

    for category_name, products in db_data.items():

        excel_products = excel_data.get(normalize_category(category_name), {})
        
        for product in products.values():

            norm_name = product["normalized_name"]


            excel_product = excel_products.get(norm_name)

            if not excel_product:
                continue

            for point_name, excel_value in excel_product.items():
                
                point_id = point_name_to_id.get(point_name)
                print(f"Для найденной точки {point_name}, присвоим point_id = {point_id} и значение {excel_value}")
                if not point_id:
                    continue

                # -------------------------
                # КЛЮЧЕВАЯ ЛОГИКА
                # -------------------------
                db_value = product["points"].get(point_id)

                if db_value is None or db_value == 0:
                    product["points"][point_id] = Decimal(str(excel_value))
    return db_data




def safe_decimal(val):

    if val is None:
        return Decimal(0)

    # если уже число
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))

    # приводим к строке
    val = str(val).strip()

    # пустые значения
    if val in ("", "-", "—"):
        return Decimal(0)

    # заменяем запятую на точку
    val = val.replace(",", ".")

    try:
        return Decimal(val)
    except  decimal.InvalidOperation:
        return Decimal(0)